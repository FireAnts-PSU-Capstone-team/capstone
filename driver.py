import collections
import os
import sys
import time
import pandas as pd
import numpy as np
import psycopg2
from openpyxl import load_workbook
from pandas.io.json import json_normalize

from db import connection as c
from models.IntakeRow import ColNames, intake_headers
from query_parser import QueryParser, RequestParseException
from validation import validate_dataframe

test_file = 'resources/sample.xlsx'
primary_table = 'intake'
db_tables = ['intake', 'txn_history', 'archive', 'metadata', 'violations', 'records']
metadata_table = 'metadata'
connection_error_msg = 'The connection to the database is closed and cannot be opened. Verify DB server is up.'


# TODO: refactor to remove duplicated code
is_connected = False
wait_time = 0
while not is_connected:
    try:
        pgSqlCur, pgSqlConn = c.pg_connect()
        is_connected = True
    except:
        time.sleep(1)
        wait_time += 1
        sys.stdout.write(f'\rConnecting to DB ... {wait_time}')


def reconnectDB():
    """
    Function to reconnect to the database if connection is closed for some reason
    Args:   None
    Return (bool): True if connection successful, else False
    """
    wait_time = 0
    global pgSqlConn, pgSqlCur
    global is_connected
    is_connected = False
    while not is_connected:
        try:
            pgSqlCur, pgSqlConn = c.pg_connect()
            is_connected = True
            return is_connected
        except:
            time.sleep(1)
            wait_time += 1
            sys.stdout.write(f'\rConnecting to DB ... {wait_time}')
            if wait_time >= 30:
                return False
    return True


def check_conn():
    """
    Function to check the connection to the database and reconnect if closed
    Args:  None
    Return (bool): True if connected, False if connection could not be established
    """
    if not is_connected:
        return reconnectDB()
    # check to make sure that the connection is open and active
    try:
        pgSqlCur.execute('SELECT 1')
        return True
    except:
        return reconnectDB()


def sql_except(err):
    """
    Function to print out the error message generated from the exception
    Args:
        err - the error message generated
    Returns None
    """
    # roll back the last sql command
    pgSqlCur.execute("ROLLBACK")
    # get the details for exception
    err_type, err_obj, traceback = sys.exc_info()

    # print the connect() error
    sys.stderr.write(f"\npsycopg2 ERROR: {err}")


def fmt(s):
    """
    Formats an input element for SQL-friendly injection. Translates None to "NULL", quotes strings, and stringifies
    non-textual arguments.
    Args:
        s: the input element
    Returns (str): a SQL-friendly string representation
    """
    if s is None or str(s).lower() == 'nan' or str(s) == '':
        s = "NULL"
    else:
        if type(s) is str:
            s = "'" + str(s).replace('"', '').replace("'", "") + "'"  # strip quotation marks
        else:
            s = str(s)
    return s


def dump_tables():
    """
    Displays the contents of the tables in the database.
    Returns: None
    """
    # check to make sure that the connection is open and active
    if not check_conn():
        return connection_error_msg

    print("Test:\n-------------------------------")
    try:
        pgSqlCur.execute(f"select * from {primary_table}")
        rows = pgSqlCur.fetchall()
        for row in rows:
            print(row)
    except Exception as err:
        sql_except(err)

    print("\nMetadata:\n-------------------------------")
    try:
        pgSqlCur.execute("select * from metadata")
        rows = pgSqlCur.fetchall()
        for row in rows:
            print(row)
    except Exception as err:
        sql_except(err)


def table_exists(cur, table):
    """
    Checks whether the named table exists.
    Args:
        cur ({}): the Postgres cursor
        table (str): the table to validate
    Returns (bool): whether the table was found
    """
    # check to make sure that the connection is open and active
    if not check_conn():
        return connection_error_msg
    try:
        cur.execute("select exists(select relname from pg_class where relname='" + table + "')")
        exists = cur.fetchone()[0]

    except psycopg2.Error:
        exists = False

    return exists
        

class InvalidTableException(Exception):
    """
    Thrown when the specified table does not exist.
    """
    pass


class InvalidRowException(Exception):
    """
    Thrown when the specified row is invalid.
    """
    pass


def filter_table(request_body):
    """
    Return a JSON object representing the requested data from the table.
    Built to take a JSON request, which is parsed in order to construct a SQL query; returns the result of that query.
    The query must comply to a schema outlined in the Swagger file.
    Args:
        request_body ({}): a JSON object, which must conform to a defined schema and is parsed to build the query
    Returns:
         query (str): the query string passed to the database
         response ({}): the retrieved data
         status (int): the HTTP status code of the response
    """
    try:
        table_names = get_table_list()
        qp = QueryParser(table_names)
        query = qp.build_query(request_body)
    except RequestParseException as e:
        return 'JSON could not be parsed', e.msg, 400
    try:
        pgSqlCur.execute(query)
        return query, pgSqlCur.fetchall(), 200
    except psycopg2.Error as err:
        sql_except(err)
        return None, str(err), 400


def get_table(table_name, columns):
    """
    Return a JSON-like format of table data.
    Args:
        table_name (str): the table to fetch
        columns ([str]): a list of columns to include in the results
    Returns ([str]): an object-notated dump of the table
    """
    result = []
    column_names = []

    # check to make sure that the connection is open and active
    if not check_conn():
        result.append(connection_error_msg)
        return result

    if not table_exists(pgSqlCur, table_name) or table_name not in db_tables:
        raise InvalidTableException

    try:
        pgSqlCur.execute(f"select * from {table_name}")
        rows = pgSqlCur.fetchall()

        for col in pgSqlCur.description:
            column_names.append(col.name)

        for row in rows:
            a_row = {}
            i = 0
            for col in column_names:
                if columns:
                    if col in columns:
                        a_row[col] = row[i]
                else:
                    a_row[col] = row[i]
                i += 1

            result.append(a_row)
            
    except Exception as err:
        sql_except(err)

    return result


def read_metadata(f):
    """
    Collects metadata about a spreadsheet to be consumed.
    Args:
        f (str): the filename of the spreadsheet
    Returns (dict): the metadata collection
    """
    headerMatches = 0
    data = {}
    workbook = load_workbook(f)
    file_data = workbook.properties.__dict__
    sheet_data = workbook.worksheets[0]
    os_data = os.stat(f)

    data['filename'] = fmt(os.path.basename(f))
    data['creator'] = fmt(file_data.get('creator'))
    data['size'] = os_data.st_size
    data['created'] = fmt(file_data.get('created').strftime('%Y-%m-%d %H:%M:%S+08'))
    data['modified'] = fmt(file_data.get('modified').strftime('%Y-%m-%d %H:%M:%S+08'))
    data['lastModifiedBy'] = fmt(file_data.get('lastModifiedBy'))
    data['title'] = fmt(file_data.get('title'))
    data['columns'] = sheet_data.max_column

    # adjust row count to account for header row, if necessary
    headerMatches = 0
    for cell in sheet_data[1]:
        if cell.value in intake_headers:
            headerMatches += 1
    if headerMatches == len(intake_headers):
        data['rows'] = sheet_data.max_row - 1
    else:
        data['rows'] = sheet_data.max_row

    return data


def write_info_data(df):
    """
    Write data from spreadsheet to the information table.
    Args:
        df (dataframe): data from spreadsheet
    Returns: dict of data writing info
    """
    # check if the connection is alive
    if not check_conn():
        return {'failure': True, 'message': connection_error_msg}
    failed_rows = []
    success_count = 0
    row_array = np.ndenumerate(df.values).iter.base
    total_count = len(row_array)
    for row in row_array:
        print(row)
        try:
            re, failed_row = insert_row(primary_table, row, True)  # TODO: need to replace so we can name a table
            if re == 1:
                success_count += 1
            else:
                failed_rows.append(failed_row)
        except:
            failed_rows.append(row)

    return {
        'insertions_attempted': total_count,
        'insertions_successful': success_count,
        'insertions_failed': failed_rows
    }


def write_metadata(metadata):
    """
    Write metadata of Excel file into metadata table.
    Args:
        metadata (dict): the metadata dictionary
    Returns: None
    """
    cmd = "INSERT INTO {}(filename, creator, size, created_date, last_modified_date, last_modified_by, title, rows, columns) " \
        "VALUES(" + "{} " + ", {}" * 8 + ") ON CONFLICT DO NOTHING"

    # check to make sure that the connection is open and active
    if not check_conn():
        return connection_error_msg

    try:
        pgSqlCur.execute(cmd.format(metadata_table, metadata['filename'], metadata['creator'], metadata['size'],
                                    metadata['created'], metadata['modified'], metadata['lastModifiedBy'],
                                    metadata['title'], metadata['rows'], metadata['columns']))
        pgSqlConn.commit()

    except Exception as err:
        sql_except(err)


def row_number_exists(cur, row_number, table=primary_table):
    """
    Checks whether the row number already exists in a table.
    Args:
        cur ({}): the Postgres cursor
        row_number (int): the row number to validate
        table (str): the table to validate, default is primary_table
    Returns (bool): whether the row number already exists
    """
    try:
        cur.execute("SELECT EXISTS(SELECT * FROM " + table + " WHERE row=" + str(row_number) + ")")
        exists = cur.fetchone()[0]

    except psycopg2.Error:
        exists = False

    return exists


def get_table_list():
    """
    Gets the database's active tables.
    Returns [str]: list of table names
    """
    try:
        pgSqlCur.execute("""
        SELECT table_name 
        FROM information_schema.tables
        WHERE table_name 
        NOT LIKE 'pg_%'
            AND table_schema='public'; 
        """)
        return str([x[0] for x in pgSqlCur.fetchall()])
    except psycopg2.Error as err:
        sql_except(err)


def validate_row(json_item):
    """
    Preps a JSON input row and passes it to the data validator. Returns the validator's response.
    Args:
        json_item ({}): input JSON
    Returns ((bool, str)): <whether row is valid>, <error message>
    """
    # If the incoming json object doesn't have a row associated with it, we add a temporary one for validation
    if 'row' not in json_item:
        json_item = collections.OrderedDict(json_item)
        json_item.update({'row': 999})
        json_item.move_to_end('row', last=False)
    df = json_normalize(json_item)
    return validate_dataframe(df)


def insert_row(table, row, checked=False):
    """
    Insert an array of values into the specified table.
    Args:
        table (str): name of table to insert into
        row ([]): row of values to insert, default to false,
        checked (bool): flag that connection to DB has already been checked by calling function
    Returns: (bool, dict) a bool indicate whether insertion is successful, a dict of failed row info
    """
    # Check flag for multi row insert, if false check to make sure that the connection is open and active
    if not checked:
        if not check_conn():
            return 0, connection_error_msg

    cmd = f"INSERT INTO {table} VALUES ("

    # Determine whether to insert at a specific row number or use default
    if row[0] is not None:
        if row_number_exists(pgSqlCur, int(row[0])):
            failed_row = {
                'submission_date': row[1],
                'entity': row[2],
                'dba': row[3],
                'message': f'Row number {row[0]} already taken.'
            }
            return 0, failed_row
        else:
            cmd += str(row[0])
    else:
        cmd += "DEFAULT"

    for i in range(1, len(row)):
        cmd += f", {fmt(row[i])}"
    cmd += ")"
    try:
        pgSqlCur.execute(cmd)
        if not checked:
            pgSqlConn.commit()
        if pgSqlCur.rowcount == 1:
            return 1, None
        else:
            raise psycopg2.Error
    except Exception as err:
        sql_except(err)
        failed_row = {
            'submission_date': row[ColNames.SUBMISSION_DATE.value],
            'entity': row[ColNames.ENTITY.value],
            'dba': row[ColNames.DBA.value],
            'mrl': row[ColNames.MRL.value]
        }
        return 0, failed_row


def process_file(f):
    """
    Read an Excel file; put info data into info table, metadata into metadata table
    Args:
        f (str): filename of spreadsheet
    Returns (bool, dict): bool is successful or not, dict includes processing info 
    """
    # check to make sure that the connection is open and active
    if not check_conn():
        return 0, connection_error_msg
    else:
        # read file content
        df = pd.read_excel(f)
        # Validate data frame
        valid, error_msg = validate_dataframe(df)
        
        # Write the data to the DB
        result_obj = write_info_data(df)
        # insert metadata into metadata table
        # should add version and revision to this schema, but don't know types yet
        metadata = read_metadata(f)

        write_metadata(metadata)

        # commit execution
        pgSqlConn.commit()
        if not valid:
            result_obj['failed_rows'] = error_msg
        failed_insertions = result_obj['insertions_attempted'] - result_obj['insertions_successful']
        return failed_insertions == 0, result_obj


def delete_row(table, row_nums):
    """
    Args:
        table (str): table name to delete row from
        row_nums ([int]): the rowId(s) to delete
    Returns (bool, dict): Boolean is successful or not, dict contains processed info
    """
    success = False
    delete_info = {}
    if not check_conn():
        return 0, connection_error_msg
    else:
        # verify table is within the db
        if table not in db_tables:
            raise InvalidTableException
        # convert each row_num to digit;
        try:
            row_nums = list(map(int, row_nums))
        except ValueError:
            raise InvalidRowException
        for row in row_nums:
            if row <= 0:
                delete_info[f'Row {str(row)}'] = 'Invalid row number'
                continue
            cmd = f'DELETE FROM {table} WHERE "row" = {row};'
            try:
                pgSqlCur.execute(cmd)
                pgSqlConn.commit()
                if pgSqlCur.rowcount == 1:
                    success = True
                    delete_info[f'Row {str(row)}'] = 'Successfully deleted'
                else:
                    delete_info[f'Row {str(row)}'] = 'Failed to delete'

            except psycopg2.Error as err:
                sql_except(err)
        if success:
            return 'Deletion successful', delete_info
        else:
            return 'Some/all deletions failed', delete_info


def test_driver():
    # Pre-insert query
    print('Dump tables -------------------------------------------------')
    dump_tables()

    print('\nInsert data -------------------------------------------------')
    process_file(test_file)

    # three options for collisions:
    # 1. do nothing (discard new row; probably want to return an error to the user in this case)
    # 2. update existing record with new metadata
    # ON CONFLICT (filename)
    #       DO UPDATE
    #       SET (size, last_modified_by, modified) = (EXCLUDED.size, EXCLUDED.last_modified_by, EXCLUDED.modified)
    # 3. add entirely new record
    # would need to return status of insertion, then insert a new row. Would need an incrementing column, like 'version'

    # unfortunately, it looks like 'creator' and 'created' are both fabricated by openpyxl, so we may need to find
    # a different way to capture that data if we want to keep them

    # Post insert query, verify data is inserted
    print('\nDump tables -------------------------------------------------')
    dump_tables()

    # close the db connection
    print("Closing connection to database.")
    c.pg_disconnect(pgSqlCur, pgSqlConn)


if __name__ == '__main__':
    test_driver()
    print(get_table(primary_table, None))
