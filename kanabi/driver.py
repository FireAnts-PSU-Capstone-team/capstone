import collections
import json
import os
import sys
import time

import numpy as np
import pandas as pd
import psycopg2

from openpyxl import load_workbook

import kanabi.db.connection as c
from kanabi.models.IntakeRow import ColNames, intake_headers, IntakeRow
from kanabi.query_parser import QueryParser, RequestParseException
from kanabi.validation.intake_validation import validate_intake
from .user import User

test_file = 'resources/sample.xlsx'
primary_table = 'intake'
metadata_table = 'metadata'
connection_error_msg = 'The connection to the database is closed and cannot be opened. Verify DB server is up.'
login_required_msg = 'Must be logged in to perform action'
row_seq = {"intake": 1, "violations": 1, "reports": 1}

# TODO: refactor to remove duplicated code
is_connected = False
wait_time = 0
while not is_connected:
    try:
        pgSqlCur, pgSqlConn = c.pg_connect()
        is_connected = True
    except:
        time.sleep(1)
        wait_time += 1
        sys.stdout.write(f'\rConnecting to DB ... {wait_time}\n')


def reconnectDB():
    """
    Function to reconnect to the database if connection is closed for some reason
    Args:   None
    Return (bool): True if connection successful, else False
    """
    wait_time = 0
    global pgSqlConn, pgSqlCur
    global is_connected
    is_connected = False
    while not is_connected:
        try:
            pgSqlCur, pgSqlConn = c.pg_connect()
            is_connected = True
            return is_connected
        except:
            time.sleep(1)
            wait_time += 1
            sys.stdout.write(f'\rConnecting to DB ... {wait_time}\n')
            if wait_time >= 30:
                return False
    return True


def table_rows_to_dicts(rows, cur, columns=None):
    """
    Constructs a dict from rows of a table.
    Args:
        rows {__iter__}: one or many rows to parse
        columns [str]: list of column names to include in results
    Returns [dict]: list of results
    """
    result = []
    column_names = []

    for col in cur.description:
        column_names.append(col.name)

    for row in rows:
        a_row = {}
        i = 0
        for col in column_names:
            if columns:
                if col in columns:
                    a_row[col] = row[i]
            else:
                a_row[col] = row[i]
            i += 1

        result.append(a_row)

    return result


def get_table_list():
    """
    Gets the database's active tables.
    Returns [str]: list of table names
    """
    try:
        pgSqlCur.execute("""
        SELECT table_name 
        FROM information_schema.tables
        WHERE table_name 
        NOT LIKE 'pg_%'
            AND table_schema='public'; 
        """)
        return str([x[0] for x in pgSqlCur.fetchall()])
    except psycopg2.Error as err:
        sql_except(err)


db_tables = get_table_list()


def check_conn():
    """
    Function to check the connection to the database and reconnect if closed
    Args:  None
    Return (bool): True if connected, False if connection could not be established
    """
    if not is_connected:
        return reconnectDB()
    # check to make sure that the connection is open and active
    try:
        pgSqlCur.execute('SELECT 1')
        return True
    except psycopg2.Error:
        return reconnectDB()


def user_cursor(user):
    """
    Creates a new Postgres connection cursor associated with the provided user.
    Args:
        user (User): associated user
    Returns (cursor, conn): Postgres connection and cursor
    """
    try:
        conn = psycopg2.connect(sslmode="require", dbname="kanabi", user=user.email, password=user.password,
                                host="db")
        cur = conn.cursor()
        conn.commit()
    except psycopg2.DatabaseError:
        return None, None
    return cur, conn


def sql_except(err):
    """
    Function to print out the error message generated from the exception.
    Args:
        err (psycopg2.Error) - the error message generated
    Returns: None
    """
    # roll back the last sql command
    pgSqlCur.execute("ROLLBACK")
    # get the details for exception
    # err_type, err_obj, traceback = sys.exc_info()

    # print the connect() error
    sys.stderr.write(f"\npsycopg2 ERROR: {err}")


def fmt(s):
    """
    Formats an input element for SQL-friendly injection. Translates None to "NULL", quotes strings, and stringifies
    non-textual arguments.
    Args:
        s (any): the input element
    Returns (str): a SQL-friendly string representation
    """
    if s is None or str(s) == '':
        s = "NULL"
    else:
        if type(s) is str:
            s = "'" + str(s).replace('"', '').replace("'", "") + "'"  # strip quotation marks
        elif str(s).lower() == 'nan' or str(s).lower() == 'nat':  # s MUST not be str
            s = "NULL"
        else:
            s = "'" + str(s) + "'"
    return s


def table_exists(cur, table):
    """
    Checks whether the named table exists.
    Args:
        cur ({}): the Postgres cursor
        table (str): the table to validate
    Returns (bool): whether the table was found
    """
    # check to make sure that the connection is open and active
    if not check_conn():
        return connection_error_msg
    try:
        cur.execute("select exists(select relname from pg_class where relname='" + table + "')")
        exists = cur.fetchone()[0]

    except psycopg2.Error:
        exists = False

    return exists


class InvalidTableException(Exception):
    """
    Thrown when the specified table does not exist.
    """
    pass


class InvalidRowException(Exception):
    """
    Thrown when the specified row is invalid.
    """
    pass


def filter_table(request_body: json, user: User) -> (str, json, int):
    """
    Return a JSON object representing the requested data from the table.
    Built to take a JSON request, which is parsed in order to construct a SQL query; returns the result of that query.
    The query must comply to a schema outlined in the Swagger file.
    Args:
        request_body ({}): a JSON object, which must conform to a defined schema and is parsed to build the query
        user: a User object holding the information for user making function call
    Returns:
         query (str): the query string passed to the database
         response ({}): the retrieved data
         status (int): the HTTP status code of the response
    """
    if not user.is_authenticated:
        return None, login_required_msg, 400
    cur, conn = user_cursor(user)
    if cur is None:
        return None, connection_error_msg, 500
    try:
        qp = QueryParser(db_tables)
        query = qp.build_query(request_body)
    except RequestParseException as e:
        return 'JSON could not be parsed', e.msg, 400
    try:
        # get our query results
        cur.execute(query)
        results = cur.fetchall()
        # get our column names
        table = request_body['table']
        col_names = request_body.get('columns')
        if col_names is None:
            if table == 'intake':
                col_names = IntakeRow.__slots__
            else:
                return None, f"Querying table {table} is not supported", 400
        # we need an array of objects where the keys are column names
        ret = []
        # results are in arrays, ordered by table index
        num_results = len([x for x in results])
        for row_num in range(num_results):
            row_result = {}
            for i, n in enumerate(col_names):
                row_result[n] = results[row_num][i]
            ret.append(row_result)
        return query, ret, 200
    except psycopg2.Error as err:
        sql_except(err)
        return None, str(err), 400


def get_table(table_name, columns, user):
    """
    Return a JSON-like format of table data.
    Args:
        table_name (str): the table to fetch
        columns ([str]): a list of columns to include in the results
        user (User): object holding user info making the function call
    Returns ([str]): an object-notated dump of the table
    """
    if not user.is_authenticated:
        return 'Must be logged in to perform action'
    result = []
    cur, conn = user_cursor(user)
    if cur is None:
        return None

    if not table_exists(cur, table_name) or table_name not in db_tables:
        raise InvalidTableException

    try:
        cur.execute(f"select * from {table_name}")
        rows = cur.fetchall()

        result = table_rows_to_dicts(rows, cur, columns)

    except Exception as err:
        sql_except(err)

    return result


def read_metadata(f):
    """
    Collects metadata about a spreadsheet to be consumed.
    Args:
        f (str): the filename of the spreadsheet
    Returns (dict): the metadata collection
    """
    data = {}
    workbook = load_workbook(f)
    file_data = workbook.properties.__dict__
    sheet_data = workbook.worksheets[0]
    os_data = os.stat(f)

    data['filename'] = fmt(os.path.basename(f))
    data['creator'] = fmt(file_data.get('creator'))
    data['size'] = os_data.st_size
    data['created'] = fmt(file_data.get('created').strftime('%Y-%m-%d %H:%M:%S+08'))
    data['modified'] = fmt(file_data.get('modified').strftime('%Y-%m-%d %H:%M:%S+08'))
    data['lastModifiedBy'] = fmt(file_data.get('lastModifiedBy'))
    data['title'] = fmt(file_data.get('title'))
    data['columns'] = sheet_data.max_column

    # adjust row count to account for header row, if necessary
    headerMatches = 0
    for cell in sheet_data[1]:
        if cell.value in intake_headers:
            headerMatches += 1
    if headerMatches == len(intake_headers):
        data['rows'] = sheet_data.max_row - 1
    else:
        data['rows'] = sheet_data.max_row

    return data


def write_info_data(df, table, user):
    """
    Write data from spreadsheet to the named table.
    Args:
        table (str): name of target table
        df (pd.DataFrame): data from spreadsheet
        user (User): User object holding info on user making funciton call
    Returns (dict): status report
    """
    # check if the connection is alive
    if not user.is_authenticated:
        return False, 'Must be logged in to perform action', 404
    failed_rows = []
    success_count = 0
    row_array = np.ndenumerate(df.values).iter.base
    total_count = len(row_array)
    for row in row_array:
        try:
            re, failed_row = insert_row(table, row, user)
            if re == 1:
                success_count += 1
            else:
                failed_rows.append(failed_row)
        except psycopg2.Error:
            failed_rows.append(row)

    return {
        'insertions_attempted': total_count,
        'insertions_successful': success_count,
        'insertions_failed': failed_rows
    }


def write_metadata(metadata, user):
    """
    Write metadata of Excel file into metadata table.
    Args:
        metadata (dict): the metadata dictionary
        user (User): User obj holding info of user making func call
    Returns: None
    """
    if not user.is_authenticated:
        return False, 'Must be logged in to perform action', 404
    cmd = "INSERT INTO {}(filename, creator, size, created_date, last_modified_date, last_modified_by, title, rows, columns) " \
          "VALUES(" + "{} " + ", {}" * 8 + ") ON CONFLICT DO NOTHING"

    cur, conn = user_cursor(user)

    try:
        cur.execute(cmd.format(metadata_table, metadata['filename'], metadata['creator'], metadata['size'],
                               metadata['created'], metadata['modified'], metadata['lastModifiedBy'],
                               metadata['title'], metadata['rows'], metadata['columns']))
        conn.commit()

    except psycopg2.Error as err:
        sql_except(err)


def row_number_exists(cur, row_number, table=primary_table):
    """
    Checks whether the row number already exists in a table.
    Args:
        cur ({}): the Postgres cursor
        row_number (int): the row number to validate
        table (str): the table to validate, default is primary_table
    Returns (bool): whether the row number already exists
    """
    try:
        cur.execute("SELECT EXISTS(SELECT * FROM " + table + " WHERE row=" + str(row_number) + ")")
        exists = cur.fetchone()[0]

    except psycopg2.Error:
        exists = False

    return exists


def validate_row(json_item, table):
    """
    Preps a JSON input row and passes it to the data validator. Returns the validator's response.
    Args:
        json_item ({}): input JSON
        table (str): table name
    Returns ((bool, str)): <whether row is valid>, <error message>
    """
    # If the incoming json object doesn't have a row associated with it, we add a temporary one for validation
    if 'row' not in json_item:
        json_item = collections.OrderedDict(json_item)
        json_item.update({'row': 999})
        json_item.move_to_end('row', last=False)
    df = pd.json_normalize(json_item)
    if table == 'intake':
        return validate_intake(df)
    # add more entries here once other tables have validators
    else:
        raise InvalidTableException


def insert_row(table, row, user):
    """
    Insert an array of values into the specified table.
    Args:
        table (str): name of table to insert into
        row ([]): row of values to insert, default to false,
        user (User): User obj that holds info on user making function call
    Returns: (bool, dict) a bool indicate whether insertion is successful, a dict of failed row info
    """
    if not user.is_authenticated:
        return False, {'Message': login_required_msg}
    cur, conn = user_cursor(user)
    global row_seq
    row_temp = row_seq[table]

    cmd = f"INSERT INTO {table} VALUES ("
    try:

        # Determine whether to insert at a specific row number or use default
        if row[0] is not None and isinstance(row[0], int):
            if row_number_exists(pgSqlCur, int(row[0]), table):
                failed_row = {
                    'row': row[0],
                    'message': f'Row number {row[0]} already taken.'
                }
                return False, failed_row
            else:
                cmd += str(row[0])
        else:
            # Loop through and update row seq to first available spot
            while row_number_exists(cur, row_temp, table):
                row_temp += 1
            cmd += f"{row_temp}"

            # if first column is not row#, then almost this is the title
            # after add a row#, add this first column as string
            if not isinstance(row[0], int) and row[0] is not None:
                cmd += "," + fmt(row[0])

        for i in range(1, len(row)):
            cmd += f", {fmt(row[i])}"
        cmd += ")"
        try:
            cur.execute(cmd)
            conn.commit()
            if cur.rowcount == 1:
                row_seq[table] = row_temp
                return True, None
            else:
                raise psycopg2.Error
        except psycopg2.Error as err:
            sql_except(err)
            if table == 'intake':
                failed_row = {
                    'mrl': row[ColNames.MRL.value]
                }
            else:
                failed_row = {
                    # TODO: once other tables have unique IDs, use them here
                    'row': row[0]
                }
            return False, failed_row
    except:
        return False, connection_error_msg


def process_file(table, file, user):
    """
    Read an Excel file; put info data into info table, metadata into metadata table
    Args:
        table (str): table into which to insert
        file (str): filename of spreadsheet
        user (User): User obj holding info on user making func call
    Returns (bool, dict): bool is successful or not, dict includes processing info
    """
    if not user.is_authenticated:
        return False, {'Message': login_required_msg}
    if table not in db_tables:
        raise InvalidTableException

    # read file content
    df = pd.read_excel(file)

    if table == 'intake':
        # Validate data frame
        valid, error_msg = validate_intake(df)
    else:
        # not validating other table than primary table for now
        valid = True
        error_msg = None

    # Write the data to the DB
    result_obj = write_info_data(df, table, user)
    # insert metadata into metadata table
    # should add version and revision to this schema, but don't know types yet
    metadata = read_metadata(file)

    write_metadata(metadata, user)

    if not valid:
        result_obj['failed_rows'] = error_msg
    failed_insertions = result_obj['insertions_attempted'] - result_obj['insertions_successful']
    return failed_insertions == 0, result_obj


def delete_row(table, row_nums, user):
    """
    Args:
        table (str): table name to delete row from
        row_nums ([int]): the row id(s) to delete
        user (User): User obj holding info on user making func call
    Returns (bool, dict): Boolean is successful or not, dict contains processed info
    """
    if not user.is_authenticated:
        return False, login_required_msg
    cur, conn = user_cursor(user)
    if cur is None:
        return False, connection_error_msg
    success = False
    delete_info = {}
    # verify table is within the db
    if table not in db_tables:
        raise InvalidTableException
    # convert each row_num to digit;
    try:
        row_nums = list(map(int, row_nums))
    except ValueError:
        raise InvalidRowException
    for row in row_nums:
        if row <= 0:
            delete_info[f'Row {str(row)}'] = 'Invalid row number'
            continue
        cmd = f'DELETE FROM {table} WHERE "row" = {row};'
        try:
            cur.execute(cmd)
            conn.commit()
            if cur.rowcount == 1:
                success = True
                delete_info[f'Row {str(row)}'] = 'Successfully deleted'
            else:
                delete_info[f'Row {str(row)}'] = 'Failed to delete'

        except psycopg2.Error as err:
            sql_except(err)
    if success:
        return 'Deletion successful', delete_info
    else:
        return 'Some/all deletions failed', delete_info


def update_table(table, row, update_columns, user):
    """
    Update one row (multiple columns) for a target table
    Args:
        table (str): table name
        row (str/int): row number
        update_columns (dict): obj of {column_name: new value, ... }
        user (User): User obj holding info for user making func call
    Returns (bool, str): bool is successful or not, str includes processing info
    """
    cur, conn = user_cursor(user)
    if cur is None:
        return False, connection_error_msg
    col_str = ''
    arg_str = ''
    exe_arg_str = ''
    arg_num = 1

    for key in update_columns:
        col_str += f', {key} = ${arg_num}'
        if isinstance(update_columns[key], int):
            arg_str += ',integer'
            exe_arg_str += f",{update_columns[key]}"
        else:
            exe_arg_str += f",'{update_columns[key]}'"
            arg_str += ',text'
        arg_num += 1

    col_str = col_str[1:]
    arg_str = arg_str[1:]
    exe_arg_str = exe_arg_str[1:]

    try:
        cur.execute(f"deallocate all;\
        prepare update_table({arg_str},integer) as \
        update {table} \
        set {col_str} \
        where row = ${arg_num};")
        cur.execute(f'execute update_table({exe_arg_str},{row});')

        if cur.rowcount != 1:
            # the target row is not updated
            return False, 'Update failed, please check if the row exists'

        # validate inserted row
        cur.execute(f"select * from {table} where row = {row}")
        new_row = cur.fetchall()
        valid, error_msg = validate_intake(pd.json_normalize(table_rows_to_dicts(new_row, cur)), 1)
        if not valid:
            cur.execute("ROLLBACK")
            conn.commit()
            return False, error_msg

    except psycopg2.Error as err:
        sql_except(err)
        return False, str(err)

    # commit if no error
    conn.commit()
    return True, 'Updated successfully'


def restore_row(row_num, user):
    """
    Function to restore a row that was previously deleted from a table
    Args:
        row_num ([int]): row number(s) in the archive table of data to restore
        user (User): User obj holding info on user making func call
    Returns (bool, str):  Bool success or not, str contains process info
    """
    if not user.is_authenticated:
        return False, login_required_msg
    cur, conn = user_cursor(user)
    if cur is None:
        return False, connection_error_msg
    restore_info = {}
    try:
        row_num = list(map(int, row_num))
    except ValueError:
        raise InvalidRowException
    # get the archive row to be restored
    try:
        success = []
        for row in row_num:
            cmd = f'SELECT restore_row({row});'
            cur.execute(cmd)
            success = cur.fetchone()
            if success[0]:
                restore_info[f'Row {str(row)}'] = 'Successfully restored'
                conn.commit()
            else:
                restore_info[f'Row {str(row)}'] = 'Failed to restore'
        return success[0], restore_info

    except psycopg2.IntegrityError:
        return False, "Can't restore the row. Duplicate row id, MRL, or receipt num."

    except psycopg2.Error as err:
        sql_except(err)
        return False, str(err)


def create_db_user(name, password, admin):
    """
    Function to create the new user in the database when they are created on webserver
    Args:
        name (str): username for new user
        password (str): password for new user
        admin (bool): flag if new user should be added to admin group
    """
    try:
        if admin:
            cmd = f'CREATE USER "{name}" WITH PASSWORD \'{password}\' IN GROUP adminaccess;'
        else:
            cmd = f'CREATE USER "{name}" WITH PASSWORD \'{password}\' IN GROUP writeaccess;'
        pgSqlCur.execute(cmd)
        if pgSqlCur.statusmessage == "CREATE ROLE":
            pgSqlConn.commit()
            return 1, f'User {name} created in the postgresDB'
        else:
            return 0, f'User {name} could not be added to the postgresDB'

    except psycopg2.Error as err:
        sql_except(err)
        return 0, str(err)
